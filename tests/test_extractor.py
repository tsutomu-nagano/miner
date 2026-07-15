from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_metadata_extractor.cli import main
from excel_metadata_extractor import extract_metadata
from excel_metadata_extractor.extractor import ExcelMetadataError


def make_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "Name"
    sheet["B1"] = "Value"
    sheet["A2"] = "Alpha"
    sheet["B2"] = 10
    sheet["A3"] = "Beta"
    sheet["B3"] = 20
    sheet["B4"] = "=SUM(B2:B3)"
    sheet["A5"] = "Merged"
    sheet["A5"].comment = Comment("sample comment", "Tester")
    sheet["A6"] = "OpenAI"
    sheet["A6"].hyperlink = "https://openai.com"
    sheet.merge_cells("A5:B5")

    table = Table(displayName="DataTable", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.create_sheet("Hidden").sheet_state = "hidden"
    workbook.save(path)


def make_statistical_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "宗教法人総括表"
    sheet["A2"] = "第1表__宗教法人数総括表"
    sheet["I4"] = "2024年12月31日現在"
    sheet["B5"] = "区分"
    sheet["C5"] = "包括宗教法人"
    sheet["D5"] = "単位宗教法人"
    sheet["I5"] = "合計"
    sheet["A8"] = "所属"
    sheet["B8"] = "系統"
    for cell in ["C8", "D8", "I8"]:
        sheet[cell] = "法人"
    rows = [
        ("文部科学大臣所轄", "神道系", 118, 22, 209),
        ("文部科学大臣所轄", "小計", 363, 300, 1170),
        ("都道府県知事所轄", "神道系", 5, 81541, 83928),
        ("全体", "合計", 387, 170195, 178537),
    ]
    for row_index, row in enumerate(rows, start=9):
        sheet.cell(row=row_index, column=1, value=row[0])
        sheet.cell(row=row_index, column=2, value=row[1])
        sheet.cell(row=row_index, column=3, value=row[2])
        sheet.cell(row=row_index, column=4, value=row[3])
        sheet.cell(row=row_index, column=9, value=row[4])
    workbook.save(path)


def test_extract_metadata(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    make_workbook(workbook_path)

    metadata = extract_metadata(workbook_path)

    assert metadata["file"]["name"] == "sample.xlsx"
    assert metadata["workbook"]["sheet_count"] == 2
    assert metadata["workbook"]["sheet_names"] == ["Data", "Hidden"]

    data_sheet = metadata["sheets"][0]
    assert data_sheet["name"] == "Data"
    assert data_sheet["dimensions"]["used_range"] == "A1:B6"
    assert data_sheet["tables"][0]["name"] == "DataTable"
    assert data_sheet["merged_cells"] == ["A5:B5"]
    assert data_sheet["cells"]["formula_count"] == 1
    assert data_sheet["cells"]["formulas"][0] == {
        "cell": "B4",
        "formula": "=SUM(B2:B3)",
    }
    assert data_sheet["cells"]["comment_count"] == 1
    assert data_sheet["cells"]["hyperlink_count"] == 1


def test_extract_statistical_metadata(tmp_path: Path) -> None:
    workbook_path = tmp_path / "statistical.xlsx"
    make_statistical_workbook(workbook_path)

    metadata = extract_metadata(workbook_path)
    statistical = metadata["statistical_metadata"][0]

    assert statistical["title"] == "第1表 宗教法人数総括表"
    assert statistical["time_axis"] == [
        {"cell": "I4", "value": "2024年12月31日現在"}
    ]
    assert statistical["data_region"]["range"] == "C9:I12"
    assert statistical["classification_items"][0]["name"] == "所属"
    assert statistical["classification_items"][1]["name"] == "系統"
    assert statistical["classification_items"][2]["name"] == "表頭分類1"
    assert statistical["classification_items"][2]["values"] == [
        "包括宗教法人",
        "単位宗教法人",
        "合計",
    ]
    assert "total_values" not in statistical["classification_items"][1]
    assert "合計" in statistical["classification_items"][1]["values"]
    assert statistical["aggregation_items"] == [
        {
            "name": "宗教法人数",
            "axis": "value",
            "cell_range": "C9:I12",
            "unit": "法人",
            "value_columns": ["C", "D", "E", "F", "G", "H", "I"],
        }
    ]
    assert statistical["value_range"] == {
        "min": 5.0,
        "max": 178537.0,
        "count": 12,
        "zero_count": 0,
    }


def test_cli_writes_json(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    output_path = tmp_path / "metadata.json"
    make_workbook(workbook_path)

    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "excel_metadata_extractor.cli",
            str(workbook_path),
            "--output",
            str(output_path),
            "--pretty",
        ],
        check=True,
        text=True,
        capture_output=True,
    )

    assert result.stdout == ""
    payload = json.loads(output_path.read_text(encoding="utf-8"))
    assert payload["sheets"][0]["name"] == "Data"


def test_cli_download_only_for_url(tmp_path: Path, monkeypatch, capsys) -> None:
    downloaded_path = tmp_path / "downloaded.xlsx"
    make_workbook(downloaded_path)

    def fake_download(url: str, output_dir: Path) -> Path:
        assert url == "https://example.test/download"
        assert output_dir == tmp_path
        return downloaded_path

    monkeypatch.setattr("excel_metadata_extractor.cli.download_workbook", fake_download)

    result = main(
        [
            "https://example.test/download",
            "--download-only",
            str(tmp_path),
        ]
    )

    assert result == 0
    assert capsys.readouterr().out.strip() == str(downloaded_path)


def test_extract_metadata_rejects_xls_in_extractor_container(tmp_path: Path) -> None:
    xls_path = tmp_path / "legacy.xls"
    xls_path.write_bytes(b"legacy excel placeholder")

    try:
        extract_metadata(xls_path)
    except ExcelMetadataError as exc:
        assert "miner-xls-converter" in str(exc)
    else:
        raise AssertionError(".xls should be converted before extraction")
