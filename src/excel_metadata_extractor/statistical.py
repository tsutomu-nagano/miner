from __future__ import annotations

import re
from dataclasses import dataclass
from numbers import Number
from typing import Any

from openpyxl.utils import get_column_letter


TIME_PATTERNS = [
    re.compile(r"\d{4}\s*年"),
    re.compile(r"\d{1,2}\s*月"),
    re.compile(r"\d{1,2}\s*日"),
    re.compile(r"\d{4}[-/]\d{1,2}([-/]\d{1,2})?"),
]
REGION_WORDS = {
    "全国",
    "都道府県",
    "北海道",
    "青森県",
    "岩手県",
    "宮城県",
    "秋田県",
    "山形県",
    "福島県",
    "茨城県",
    "栃木県",
    "群馬県",
    "埼玉県",
    "千葉県",
    "東京都",
    "神奈川県",
    "新潟県",
    "富山県",
    "石川県",
    "福井県",
    "山梨県",
    "長野県",
    "岐阜県",
    "静岡県",
    "愛知県",
    "三重県",
    "滋賀県",
    "京都府",
    "大阪府",
    "兵庫県",
    "奈良県",
    "和歌山県",
    "鳥取県",
    "島根県",
    "岡山県",
    "広島県",
    "山口県",
    "徳島県",
    "香川県",
    "愛媛県",
    "高知県",
    "福岡県",
    "佐賀県",
    "長崎県",
    "熊本県",
    "大分県",
    "宮崎県",
    "鹿児島県",
    "沖縄県",
}


@dataclass(frozen=True)
class DataRegion:
    first_row: int
    last_row: int
    first_col: int
    last_col: int

    @property
    def range(self) -> str:
        return (
            f"{get_column_letter(self.first_col)}{self.first_row}:"
            f"{get_column_letter(self.last_col)}{self.last_row}"
        )


def extract_statistical_metadata(sheet: Any) -> dict[str, Any]:
    values = _sheet_values(sheet)
    data_region = _detect_data_region(values)
    if data_region is None:
        return {
            "title": _detect_title(values),
            "time_axis": _detect_time_axis(values),
            "data_region": None,
            "classification_items": [],
            "aggregation_items": [],
            "regional_items": [],
            "value_range": None,
        }

    header_rows = list(range(1, data_region.first_row))
    row_header_cols = list(range(1, data_region.first_col))
    title = _detect_title(values)
    classification_items = _row_dimensions(values, data_region, row_header_cols)
    classification_items.extend(_column_classifications(values, data_region, header_rows))
    aggregation_items = _aggregation_items(title, values, data_region, header_rows)
    regional_items = _regional_items(classification_items)
    numeric_values = _numeric_values(values, data_region)

    return {
        "title": title,
        "time_axis": _detect_time_axis(values),
        "data_region": {
            "range": data_region.range,
            "first_data_row": data_region.first_row,
            "last_data_row": data_region.last_row,
            "first_value_column": get_column_letter(data_region.first_col),
            "last_value_column": get_column_letter(data_region.last_col),
        },
        "classification_items": classification_items,
        "aggregation_items": aggregation_items,
        "regional_items": regional_items,
        "value_range": _value_range(numeric_values),
    }


def _sheet_values(sheet: Any) -> list[list[Any]]:
    return [
        [cell.value for cell in row]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        )
    ]


def _detect_data_region(values: list[list[Any]]) -> DataRegion | None:
    numeric_positions = [
        (row_index, col_index)
        for row_index, row in enumerate(values, start=1)
        for col_index, value in enumerate(row, start=1)
        if _is_number(value)
    ]
    if not numeric_positions:
        return None

    row_counts: dict[int, int] = {}
    for row_index, _ in numeric_positions:
        row_counts[row_index] = row_counts.get(row_index, 0) + 1

    candidate_rows = [
        row_index
        for row_index, count in row_counts.items()
        if count >= 2 or count == max(row_counts.values())
    ]
    first_row = min(candidate_rows)
    last_row = max(candidate_rows)

    candidate_cols = [
        col_index
        for row_index, col_index in numeric_positions
        if first_row <= row_index <= last_row
    ]
    return DataRegion(
        first_row=first_row,
        last_row=last_row,
        first_col=min(candidate_cols),
        last_col=max(candidate_cols),
    )


def _detect_title(values: list[list[Any]]) -> str | None:
    for row in values:
        if any(_is_number(value) for value in row):
            break
        texts = [
            value
            for value in _text_values(row)
            if not _looks_like_time(value) and len(value) >= 3
        ]
        if len(texts) == 1:
            return texts[0]

    candidates = []
    for row in values:
        if any(_is_number(value) for value in row):
            break
        candidates.extend(
            value
            for value in _text_values(row)
            if not _looks_like_time(value) and len(value) >= 3
        )
    if not candidates:
        return None
    return max(candidates, key=len)


def _detect_time_axis(values: list[list[Any]]) -> list[dict[str, str]]:
    items = []
    for row_index, row in enumerate(values, start=1):
        for col_index, value in enumerate(row, start=1):
            text = _clean_text(value)
            if text and _looks_like_time(text):
                items.append(
                    {
                        "cell": f"{get_column_letter(col_index)}{row_index}",
                        "value": text,
                    }
                )
    return items


def _row_dimensions(
    values: list[list[Any]],
    data_region: DataRegion,
    row_header_cols: list[int],
) -> list[dict[str, Any]]:
    items = []
    for col_index in row_header_cols:
        labels = []
        for row_index in range(data_region.first_row, data_region.last_row + 1):
            value = _cell(values, row_index, col_index)
            text = _clean_text(value)
            if text:
                labels.append(text)

        unique_labels = _unique(labels)
        if not unique_labels:
            continue

        header = _nearest_header(values, data_region.first_row, col_index)
        items.append(
            {
                "name": header or f"列{get_column_letter(col_index)}",
                "axis": "row",
                "cell_range": (
                    f"{get_column_letter(col_index)}{data_region.first_row}:"
                    f"{get_column_letter(col_index)}{data_region.last_row}"
                ),
                "values": unique_labels,
            }
        )
    return items


def _column_classifications(
    values: list[list[Any]],
    data_region: DataRegion,
    header_rows: list[int],
) -> list[dict[str, Any]]:
    active_header_rows = [
        row_index
        for row_index in header_rows
        if any(
            (text := _clean_text(_cell(values, row_index, col_index)))
            and not _looks_like_time(text)
            and not _is_unit_label(text)
            for col_index in range(data_region.first_col, data_region.last_col + 1)
        )
    ]

    level_values: dict[int, dict[str, Any]] = {}
    for col_index in range(data_region.first_col, data_region.last_col + 1):
        for level_index, row_index in enumerate(active_header_rows, start=1):
            text = _clean_text(_cell(values, row_index, col_index))
            if not text or _looks_like_time(text) or _is_unit_label(text):
                continue
            level_values.setdefault(level_index, {"row": row_index, "values": []})
            level_values[level_index]["values"].append(text)

    items = []
    for level_index, level in level_values.items():
        unique_values = _unique(level["values"])
        if not unique_values:
            continue
        items.append(
            {
                "name": _column_classification_name(level_index),
                "axis": "column",
                "level": level_index,
                "cell_range": (
                    f"{get_column_letter(data_region.first_col)}{level['row']}:"
                    f"{get_column_letter(data_region.last_col)}{level['row']}"
                ),
                "values": unique_values,
            }
        )
    return items


def _aggregation_items(
    title: str | None,
    values: list[list[Any]],
    data_region: DataRegion,
    header_rows: list[int],
) -> list[dict[str, Any]]:
    unit = _detect_unit(values, data_region, header_rows)
    return [
        {
            "name": _measure_name_from_title(title) or "数値",
            "axis": "value",
            "cell_range": data_region.range,
            "unit": unit,
            "value_columns": [
                get_column_letter(col_index)
                for col_index in range(data_region.first_col, data_region.last_col + 1)
            ],
        }
    ]


def _column_classification_name(level_index: int) -> str:
    return f"表頭分類{level_index}"


def _detect_unit(
    values: list[list[Any]],
    data_region: DataRegion,
    header_rows: list[int],
) -> str | None:
    candidates = []
    for col_index in range(data_region.first_col, data_region.last_col + 1):
        for row_index in header_rows:
            text = _clean_text(_cell(values, row_index, col_index))
            if text and _is_unit_label(text):
                candidates.append(text)
    if not candidates:
        return None
    return _unique(candidates)[0]


def _measure_name_from_title(title: str | None) -> str | None:
    if not title:
        return None
    text = re.sub(r"^第\s*\d+\s*表\s*", "", title)
    text = text.replace("総括表", "").replace("一覧表", "").replace("集計表", "")
    text = text.strip()
    return text or title


def _regional_items(classification_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    regional = []
    for item in classification_items:
        values = item["values"]
        if _looks_like_region_name(item["name"]) or any(
            _looks_like_region_name(value) for value in values
        ):
            regional.append(item)
    return regional


def _numeric_values(values: list[list[Any]], data_region: DataRegion) -> list[float]:
    result = []
    for row_index in range(data_region.first_row, data_region.last_row + 1):
        for col_index in range(data_region.first_col, data_region.last_col + 1):
            value = _cell(values, row_index, col_index)
            if _is_number(value):
                result.append(float(value))
    return result


def _value_range(values: list[float]) -> dict[str, Any] | None:
    if not values:
        return None
    return {
        "min": min(values),
        "max": max(values),
        "count": len(values),
        "zero_count": sum(1 for value in values if value == 0),
    }


def _nearest_header(values: list[list[Any]], first_data_row: int, col_index: int) -> str | None:
    for row_index in range(first_data_row - 1, 0, -1):
        text = _clean_text(_cell(values, row_index, col_index))
        if text and not _looks_like_time(text):
            return text
    return None


def _cell(values: list[list[Any]], row_index: int, col_index: int) -> Any:
    try:
        return values[row_index - 1][col_index - 1]
    except IndexError:
        return None


def _text_values(row: list[Any]) -> list[str]:
    return [text for value in row if (text := _clean_text(value))]


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\n", " ").replace("__", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def _is_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, Number):
        return True
    if isinstance(value, str):
        text = value.replace(",", "").strip()
        return bool(re.fullmatch(r"-?\d+(\.\d+)?", text))
    return False


def _looks_like_time(value: str) -> bool:
    return any(pattern.search(value) for pattern in TIME_PATTERNS)


def _looks_like_region_name(value: str) -> bool:
    return any(word in value for word in REGION_WORDS)


def _is_unit_label(value: str) -> bool:
    return len(value) <= 6 and value in {"人", "世帯", "件", "法人", "円", "千円", "百万円", "年", "月", "%", "％"}


def _unique(values: list[str]) -> list[str]:
    result = []
    seen = set()
    for value in values:
        if value not in seen:
            result.append(value)
            seen.add(value)
    return result
