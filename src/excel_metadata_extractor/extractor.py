from __future__ import annotations

from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .statistical import extract_statistical_metadata


SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


class ExcelMetadataError(ValueError):
    """Raised when metadata cannot be extracted from a workbook."""


def extract_metadata(
    workbook_path: str | Path,
    *,
    include_empty_cells: bool = False,
) -> dict[str, Any]:
    path = Path(workbook_path)
    if not path.exists():
        raise ExcelMetadataError(f"workbook not found: {path}")
    if path.suffix.lower() == ".xls":
        raise ExcelMetadataError(
            ".xls files must be converted with the separate miner-xls-converter "
            "container before extraction"
        )
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        supported = ", ".join(sorted(SUPPORTED_SUFFIXES))
        raise ExcelMetadataError(
            f"unsupported workbook extension '{path.suffix}'. Supported: {supported}"
        )

    workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        return {
            "file": _file_metadata(path),
            "workbook": _workbook_metadata(workbook),
            "defined_names": _defined_names(workbook),
            "sheets": [
                _sheet_metadata(sheet, include_empty_cells=include_empty_cells)
                for sheet in workbook.worksheets
            ],
            "statistical_metadata": [
                {
                    "sheet_name": sheet.title,
                    **extract_statistical_metadata(sheet),
                }
                for sheet in workbook.worksheets
            ],
        }
    finally:
        workbook.close()


def _file_metadata(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "name": path.name,
        "path": str(path.resolve()),
        "extension": path.suffix.lower(),
        "size_bytes": stat.st_size,
        "modified_at": _to_iso(datetime.fromtimestamp(stat.st_mtime).astimezone()),
    }


def _workbook_metadata(workbook: Any) -> dict[str, Any]:
    props = workbook.properties
    return {
        "active_sheet": workbook.active.title if workbook.worksheets else None,
        "sheet_count": len(workbook.worksheets),
        "sheet_names": workbook.sheetnames,
        "calculation": {
            "mode": getattr(workbook.calculation, "calcMode", None),
            "full_calc_on_load": getattr(workbook.calculation, "fullCalcOnLoad", None),
            "force_full_calc": getattr(workbook.calculation, "forceFullCalc", None),
        },
        "properties": {
            "title": props.title,
            "subject": props.subject,
            "creator": props.creator,
            "last_modified_by": props.lastModifiedBy,
            "created": _to_iso(props.created),
            "modified": _to_iso(props.modified),
            "category": props.category,
            "keywords": props.keywords,
            "description": props.description,
            "language": props.language,
            "version": props.version,
            "revision": props.revision,
        },
    }


def _defined_names(workbook: Any) -> list[dict[str, Any]]:
    names = []
    for name, definition in workbook.defined_names.items():
        destinations = []
        try:
            destinations = [
                {"sheet": sheet, "range": cells}
                for sheet, cells in definition.destinations
            ]
        except Exception:
            destinations = []

        names.append(
            {
                "name": name,
                "value": definition.attr_text,
                "scope": definition.localSheetId,
                "hidden": definition.hidden,
                "destinations": destinations,
            }
        )
    return names


def _sheet_metadata(sheet: Any, *, include_empty_cells: bool) -> dict[str, Any]:
    dimensions = _dimensions(sheet)
    cells = _cell_summary(sheet, include_empty_cells=include_empty_cells)

    return {
        "name": sheet.title,
        "state": sheet.sheet_state,
        "type": "worksheet",
        "dimensions": dimensions,
        "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "auto_filter": sheet.auto_filter.ref,
        "tables": _tables(sheet),
        "merged_cells": [str(cell_range) for cell_range in sheet.merged_cells.ranges],
        "charts_count": len(sheet._charts),
        "images_count": len(sheet._images),
        "data_validations_count": len(sheet.data_validations.dataValidation),
        "conditional_formatting_count": len(sheet.conditional_formatting),
        "cells": cells,
        "protection": {
            "enabled": sheet.protection.sheet,
            "password_hash": sheet.protection.password,
        },
    }


def _dimensions(sheet: Any) -> dict[str, Any]:
    max_row = sheet.max_row or 0
    max_column = sheet.max_column or 0
    used_range = None
    if max_row and max_column:
        used_range = f"A1:{get_column_letter(max_column)}{max_row}"

    return {
        "used_range": used_range,
        "max_row": max_row,
        "max_column": max_column,
    }


def _tables(sheet: Any) -> list[dict[str, Any]]:
    tables = []
    for table in sheet.tables.values():
        tables.append(
            {
                "name": table.name,
                "display_name": table.displayName,
                "range": table.ref,
                "style": table.tableStyleInfo.name if table.tableStyleInfo else None,
                "show_first_column": (
                    table.tableStyleInfo.showFirstColumn
                    if table.tableStyleInfo
                    else None
                ),
                "show_last_column": (
                    table.tableStyleInfo.showLastColumn
                    if table.tableStyleInfo
                    else None
                ),
                "show_row_stripes": (
                    table.tableStyleInfo.showRowStripes
                    if table.tableStyleInfo
                    else None
                ),
            }
        )
    return tables


def _cell_summary(sheet: Any, *, include_empty_cells: bool) -> dict[str, Any]:
    data_types: Counter[str] = Counter()
    formulas: list[dict[str, str]] = []
    comments: list[dict[str, str | None]] = []
    hyperlinks: list[dict[str, str | None]] = []
    non_empty_count = 0
    scanned_count = 0

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None and not include_empty_cells:
                continue

            scanned_count += 1
            if cell.value is not None:
                non_empty_count += 1
            data_types[cell.data_type] += 1

            if cell.data_type == "f":
                formulas.append({"cell": cell.coordinate, "formula": str(cell.value)})
            if cell.comment:
                comments.append(
                    {
                        "cell": cell.coordinate,
                        "author": cell.comment.author,
                        "text": cell.comment.text,
                    }
                )
            if cell.hyperlink:
                hyperlinks.append(
                    {
                        "cell": cell.coordinate,
                        "target": cell.hyperlink.target,
                        "location": cell.hyperlink.location,
                        "display": cell.hyperlink.display,
                    }
                )

    return {
        "scanned_count": scanned_count,
        "non_empty_count": non_empty_count,
        "data_types": dict(sorted(data_types.items())),
        "formula_count": len(formulas),
        "formulas": formulas,
        "comment_count": len(comments),
        "comments": comments,
        "hyperlink_count": len(hyperlinks),
        "hyperlinks": hyperlinks,
    }


def _to_iso(value: Any) -> str | None:
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if value is None:
        return None
    return str(value)
