from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def extract_sheet_previews(
    workbook_path: str | Path,
    *,
    max_rows: int = 80,
    max_cols: int = 40,
) -> list[dict[str, Any]]:
    workbook = load_workbook(
        workbook_path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        return [
            {
                "sheet_name": sheet.title,
                "range": _preview_range(sheet.max_row or 0, sheet.max_column or 0, max_rows, max_cols),
                "truncated": (sheet.max_row or 0) > max_rows
                or (sheet.max_column or 0) > max_cols,
                "max_row": sheet.max_row or 0,
                "max_column": sheet.max_column or 0,
                "rows": [
                    [_cell_value(cell.value) for cell in row]
                    for row in sheet.iter_rows(
                        min_row=1,
                        max_row=min(sheet.max_row or 0, max_rows),
                        min_col=1,
                        max_col=min(sheet.max_column or 0, max_cols),
                    )
                ],
            }
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _preview_range(max_row: int, max_col: int, row_limit: int, col_limit: int) -> str | None:
    if max_row == 0 or max_col == 0:
        return None
    last_row = min(max_row, row_limit)
    last_col = min(max_col, col_limit)
    return f"A1:{get_column_letter(last_col)}{last_row}"


def _cell_value(value: Any) -> str | int | float | bool | None:
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    return value

